"""
═══════════════════════════════════════════════════════
Omega CivicFlow — DART File Parser
DART 공시 파일 전용 파서 (XBRL, ZIP, XLS)

XBRL ZIP → 한국어 레이블 + 재무 데이터 추출
XLS/XLSX → 테이블 텍스트 추출
═══════════════════════════════════════════════════════
"""

import io
import re
import logging
import zipfile
import tempfile
from pathlib import Path
from typing import Optional, Tuple

logger = logging.getLogger(__name__)


def extract_text_from_dart_zip(content: bytes, filename: str = "") -> str:
    """
    DART XBRL ZIP 파일에서 구조화 텍스트 추출

    ZIP 구조:
    - *.xbrl       → 메인 재무 데이터
    - *_lab-ko.xml  → 한국어 레이블 (태그→이름 매핑)
    - *_lab-en.xml  → 영어 레이블
    - *.xsd         → 스키마
    - *_def/cal/pre.xml → 관계 정의
    """
    try:
        z = zipfile.ZipFile(io.BytesIO(content))
    except (zipfile.BadZipFile, Exception) as e:
        logger.warning(f"ZIP 파일 열기 실패: {e}")
        return ""

    names = z.namelist()
    xbrl_files = [n for n in names if n.endswith('.xbrl')]
    label_ko_files = [n for n in names if 'lab-ko' in n or 'lab_' in n and '-ko' in n]
    label_en_files = [n for n in names if 'lab-en' in n or 'lab_' in n and '-en' in n]

    # 한국어 레이블 사전 구축
    labels = {}
    for lf in label_ko_files:
        try:
            labels.update(_parse_label_file(z.read(lf)))
        except Exception:
            pass
    if not labels:
        for lf in label_en_files:
            try:
                labels.update(_parse_label_file(z.read(lf)))
            except Exception:
                pass

    # XBRL 메인 파일에서 재무 데이터 추출
    all_text = []
    for xf in xbrl_files:
        try:
            data = z.read(xf).decode('utf-8', errors='replace')
            text = _parse_xbrl_data(data, labels, filename)
            if text:
                all_text.append(text)
        except Exception as e:
            logger.warning(f"XBRL 파싱 실패 ({xf}): {e}")

    # XBRL이 없으면 모든 XML/XSD에서 텍스트 추출
    if not all_text:
        for name in names:
            if name.endswith(('.xml', '.xsd', '.xbrl')):
                try:
                    data = z.read(name).decode('utf-8', errors='replace')
                    text = _extract_xml_text(data)
                    if text and len(text) > 50:
                        all_text.append(f"[{name}]\n{text}")
                except Exception:
                    pass

    return "\n\n".join(all_text) if all_text else ""


def _parse_label_file(content: bytes) -> dict:
    """레이블 XML에서 {element_id: 한국어 이름} 사전 추출"""
    from bs4 import BeautifulSoup
    labels = {}
    try:
        text = content.decode('utf-8', errors='replace')
        soup = BeautifulSoup(text, "lxml-xml")
        # XBRL 2.1 label linkbase
        for label in soup.find_all(True, attrs={"label": True}):
            lab_text = label.string
            if lab_text and lab_text.strip():
                lab_id = label.get("id", "") or label.get("label", "")
                labels[lab_id] = lab_text.strip()

        # 더 일반적인 패턴: <label ...>텍스트</label>
        for tag in soup.find_all(re.compile(r'label$', re.I)):
            if tag.string and tag.string.strip():
                # xlink:label 속성에서 ID 추출
                for attr_name in ['xlink:label', 'label', 'id']:
                    attr_val = tag.get(attr_name, "")
                    if attr_val:
                        labels[attr_val] = tag.string.strip()
                        break

        # loc → label 매핑 (labelArc를 통해)
        locs = {}
        for loc in soup.find_all(re.compile(r'loc$', re.I)):
            href = loc.get('xlink:href', '') or loc.get('href', '')
            label_attr = loc.get('xlink:label', '')
            if href and label_attr:
                # href에서 element 이름 추출: ...#ifrs-full_Revenue
                if '#' in href:
                    elem = href.split('#')[-1]
                    locs[label_attr] = elem

        # labelArc로 loc→label 연결
        for arc in soup.find_all(re.compile(r'labelArc$', re.I)):
            from_attr = arc.get('xlink:from', '')
            to_attr = arc.get('xlink:to', '')
            if from_attr in locs and to_attr in labels:
                elem_name = locs[from_attr]
                labels[elem_name] = labels[to_attr]

    except Exception as e:
        logger.debug(f"레이블 파싱 부분 실패: {e}")

    return labels


def _parse_xbrl_data(xbrl_text: str, labels: dict, filename: str = "") -> str:
    """XBRL 인스턴스 문서에서 재무 데이터 추출"""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(xbrl_text, "lxml-xml")

    entries = []
    context_dates = {}

    # 컨텍스트에서 날짜 정보 추출
    for ctx in soup.find_all(re.compile(r'context$', re.I)):
        ctx_id = ctx.get('id', '')
        period = ctx.find(re.compile(r'(instant|startDate|endDate)$', re.I))
        if period and period.string:
            context_dates[ctx_id] = period.string.strip()
        # period 태그에서 찾기
        start = ctx.find(re.compile(r'startDate$', re.I))
        end = ctx.find(re.compile(r'endDate$', re.I))
        if start and start.string and end and end.string:
            context_dates[ctx_id] = f"{start.string.strip()}~{end.string.strip()}"
        instant = ctx.find(re.compile(r'instant$', re.I))
        if instant and instant.string:
            context_dates[ctx_id] = instant.string.strip()

    # 실제 데이터 노드 추출 (숫자값이 있는 요소들)
    for tag in soup.find_all(True):
        if not tag.string or not tag.string.strip():
            continue
        text = tag.string.strip()

        # 네임스페이스 분리
        tag_name = tag.name
        if ':' in tag_name:
            ns, local = tag_name.split(':', 1)
        else:
            local = tag_name

        # 컨텍스트/단위 속성이 있는 것만 = 실제 데이터
        ctx_ref = tag.get('contextRef', '')
        unit_ref = tag.get('unitRef', '')
        decimals = tag.get('decimals', '')

        if ctx_ref or unit_ref:
            # 한국어 레이블 찾기
            korean_label = _find_label(local, labels)
            period = context_dates.get(ctx_ref, '')

            if korean_label:
                entry = f"{korean_label}: {text}"
            else:
                # 언더스코어/CamelCase를 읽기 좋게 변환
                readable = re.sub(r'([a-z])([A-Z])', r'\1 \2', local)
                readable = readable.replace('_', ' ')
                entry = f"{readable}: {text}"

            if period:
                entry += f" ({period})"
            if unit_ref and 'KRW' in unit_ref.upper():
                # 큰 숫자 포맷
                try:
                    num = float(text)
                    if abs(num) >= 1e8:
                        entry += f" [≈{num/1e8:.0f}억원]"
                except ValueError:
                    pass

            entries.append(entry)

    if not entries:
        # fallback: 모든 텍스트 노드에서 추출
        return _extract_xml_text(xbrl_text)

    # 파일명에서 회사명 추출
    header = ""
    if filename:
        company_match = re.search(r'\[(.+?)\]', filename)
        if company_match:
            header = f"회사명: {company_match.group(1)}\n"

    return header + "\n".join(entries)


def _find_label(element_name: str, labels: dict) -> str:
    """element 이름으로 한국어 레이블 검색 (퍼지 매칭)"""
    if not labels:
        return ""

    # 정확히 일치
    if element_name in labels:
        return labels[element_name]

    # 대소문자 무시 매칭
    lower = element_name.lower()
    for key, val in labels.items():
        if key.lower() == lower:
            return val
        # element_name이 key에 포함
        if lower in key.lower() or key.lower() in lower:
            return val

    return ""


def _extract_xml_text(xml_text: str) -> str:
    """일반 XML에서 의미있는 텍스트 추출 (fallback)"""
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(xml_text, "lxml-xml")
    texts = []
    for tag in soup.find_all(True):
        if tag.string and tag.string.strip():
            tag_name = tag.name.split(':')[-1] if ':' in tag.name else tag.name
            texts.append(f"{tag_name}: {tag.string.strip()}")
    return "\n".join(texts) if texts else soup.get_text(separator="\n", strip=True)


def extract_text_from_xls(content: bytes, filename: str = "") -> str:
    """
    XLS/XLSX 파일에서 텍스트 추출

    XLS (OLE2) → xlrd
    XLSX → openpyxl
    """
    text_parts = []

    # XLSX 먼저 시도
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(c for c in cells if c))
            if rows:
                text_parts.append(f"[시트: {sheet_name}]\n" + "\n".join(rows))
        wb.close()
        if text_parts:
            return "\n\n".join(text_parts)
    except Exception:
        pass

    # XLS (OLE2) — xlrd로 시도
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        for sheet in wb.sheets():
            rows = []
            for rx in range(sheet.nrows):
                cells = [str(sheet.cell_value(rx, cx)) for cx in range(sheet.ncols)]
                if any(c.strip() for c in cells):
                    rows.append(" | ".join(c for c in cells if c.strip()))
            if rows:
                text_parts.append(f"[시트: {sheet.name}]\n" + "\n".join(rows))
        if text_parts:
            return "\n\n".join(text_parts)
    except ImportError:
        logger.warning("xlrd 미설치 — pip install xlrd")
    except Exception as e:
        logger.warning(f"XLS 파싱 실패: {e}")

    # 최후의 수단: 바이너리에서 텍스트 추출 시도
    try:
        text = content.decode('utf-8', errors='replace')
        # 의미있는 텍스트만 추출
        clean = re.sub(r'[^\x20-\x7E\uAC00-\uD7A3\u3131-\u3163\u1100-\u11FF\n]', ' ', text)
        clean = re.sub(r' {3,}', '\n', clean)
        clean = "\n".join(line.strip() for line in clean.split('\n') if len(line.strip()) > 5)
        if len(clean) > 100:
            return clean
    except Exception:
        pass

    return ""
