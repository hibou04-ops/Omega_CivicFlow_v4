#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
update_deliverables.py
Updates 10 Excel deliverable files for Omega CivicFlow v4.
Applies factual corrections based on actual codebase (2026-04-16).
"""
import sys, os, io, shutil, copy
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy as shallow_copy

BASE_DIR = r'C:\Users\hibou\Downloads\새 폴더\산출물_4조_곽경훈'

# ─── Style helpers ───────────────────────────────────────────────────────
THIN = Side(style='thin')
THIN_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
FONT_10 = Font(name='맑은 고딕', size=10)
FONT_11 = Font(name='맑은 고딕', size=11)
ALIGN_CC_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_TOP_WRAP = Alignment(vertical='top', wrap_text=True)


def copy_cell_style(src_cell, dst_cell):
    """Copy style attributes from src to dst cell."""
    dst_cell.font = copy.copy(src_cell.font)
    dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.border = copy.copy(src_cell.border)
    dst_cell.fill = copy.copy(src_cell.fill)
    dst_cell.number_format = src_cell.number_format


def backup_file(filepath):
    """Create .bak backup."""
    bak = filepath + '.bak'
    if not os.path.exists(bak):
        shutil.copy2(filepath, bak)
        print(f"  [BACKUP] {os.path.basename(bak)}")


def write_row(ws, row_num, values, style_row=None):
    """Write a row of values and optionally copy style from style_row."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        if style_row:
            src = ws.cell(row=style_row, column=col_idx)
            copy_cell_style(src, cell)


def append_styled_row(ws, values, style_row=None):
    """Append a new row after max_row, copying style from style_row."""
    new_row = ws.max_row + 1
    write_row(ws, new_row, values, style_row=style_row)
    return new_row


# ═══════════════════════════════════════════════════════════════════════════
# FILE 1: 01_분석_요구사항정의서_v1.0.xlsx
# ═══════════════════════════════════════════════════════════════════════════
def update_file_01():
    fname = '01_분석_요구사항정의서_v1.0.xlsx'
    fpath = os.path.join(BASE_DIR, fname)
    print(f"\n[FILE 1] {fname}")
    backup_file(fpath)
    wb = openpyxl.load_workbook(fpath)

    # --- 표지 ---
    ws = wb['표지']
    ws.cell(row=11, column=4, value='v2.0')
    ws.cell(row=12, column=4, value='2026.04.16')
    print("  [표지] version → v2.0, date → 2026.04.16")

    # --- 개정이력 ---
    ws = wb['개정이력']
    # Existing data row is R4 (style_row=4), new row at R5
    new_vals = ['2.0', '2026.04.16',
                '소스코드 기반 전면 최신화 (RAG 챗봇, DART 검색, 패널, 회원탈퇴, 홈페이지 등 추가)',
                '곽경훈', '곽경훈']
    append_styled_row(ws, new_vals, style_row=4)
    print("  [개정이력] Added v2.0 row")

    # --- 1. 기능 요구사항 ---
    ws = wb['1. 기능 요구사항']
    style_r = 9  # last existing data row

    new_reqs = [
        ['RQ-FN-009', '핵심기능', 'Omega-Prime RAG 챗봇',
         'Omega-Prime 하이브리드 RAG 에이전트 챗봇\n- POST /panel/chat: 인증 필수, 실시간 대화\n- Tool-Based RAG (DB 쿼리) + Vector RAG (시맨틱 검색)\n- 하이브리드 검색: Vector(0.4) + BM25(0.25) + Metadata(0.35)\n- CrossEncoder 리랭킹 (bge-reranker-v2-m3-ko)\n- 4개 에이전트 오케스트레이션 (Router→Planner→Judge→Synthesizer)',
         '상', '화면설계서', '사용', '분석완료', '신규', '곽경훈 (4조)'],

        ['RQ-FN-010', '핵심기능', 'DART 공시 검색',
         'DART OpenAPI 연동 공시 검색\n- POST /panel/search: 종목명/코드 기반 공시 검색\n- GET /panel/autocomplete: 상장법인 자동완성\n- corpCode.xml 캐싱 (~80,000 법인)\n- DART 공시 URL 직접 링크 제공',
         '상', '화면설계서', '사용', '분석완료', '신규', '곽경훈 (4조)'],

        ['RQ-FN-011', '조회', '사이드 패널 (실시간 모니터링)',
         '사이드 패널 실시간 시스템 모니터링\n- GET /panel/stats: DB 집계 통계 (문서수, 분석율 등)\n- GET /panel/system-status: 서비스 헬스체크 + 경보 레벨\n- GET /panel/activity-log: 최근 활동 로그',
         '중', '화면설계서', '사용', '분석완료', '신규', '곽경훈 (4조)'],

        ['RQ-FN-012', '인증', '회원탈퇴 (2단계 이메일 인증)',
         '회원탈퇴 2단계 이메일 인증 플로우\n- POST /auth/request-withdraw: 비밀번호 + 확인문구 검증 후 인증메일 발송\n- POST /auth/confirm-withdraw: 토큰 검증 후 PII 익명화 + 비활성화\n- PIPA(개인정보보호법) 준수: 이메일/사용자명/비밀번호 해시 즉시 익명화',
         '상', '화면설계서', '사용', '분석완료', '신규', '곽경훈 (4조)'],

        ['RQ-FN-013', '조회', '홈페이지',
         '비로그인 사용자 접근 가능한 서비스 소개 홈페이지\n- /home URL (기본 라우트 / → /home 리다이렉트)\n- 서비스 소개 및 기능 안내\n- 로그인/회원가입 유도',
         '중', '화면설계서', '사용', '분석완료', '신규', '곽경훈 (4조)'],

        ['RQ-FN-014', '관리', '벡터 인덱스 관리',
         'ChromaDB 벡터 인덱스 관리\n- POST /panel/vector/rebuild: 전체 재인덱싱 (백그라운드)\n- GET /panel/vector/stats: 인덱스 현황 조회\n- BGE-M3 1024-dim 임베딩, 284K+ 벡터 청크',
         '중', '화면설계서', '사용', '분석완료', '신규', '곽경훈 (4조)'],

        ['RQ-FN-015', '조회', 'PDF 보고서 생성/다운로드',
         '분석 결과 PDF 보고서 생성 및 다운로드\n- GET /documents/download-report/{id}: PDF 다운로드\n- GET /documents/preview-report/{id}: 브라우저 미리보기\n- fpdf2 기반 자동 생성',
         '중', '화면설계서', '사용', '분석완료', '신규', '곽경훈 (4조)'],
    ]

    for req_data in new_reqs:
        append_styled_row(ws, req_data, style_row=style_r)

    print(f"  [기능 요구사항] Added {len(new_reqs)} rows (RQ-FN-009 ~ RQ-FN-015)")

    wb.save(fpath)
    print(f"  [SAVED] {fname}")


# ═══════════════════════════════════════════════════════════════════════════
# FILE 2: 02_분석_기능구조도_v1.0.xlsx
# ═══════════════════════════════════════════════════════════════════════════
def update_file_02():
    fname = '02_분석_기능구조도_v1.0.xlsx'
    fpath = os.path.join(BASE_DIR, fname)
    print(f"\n[FILE 2] {fname}")
    backup_file(fpath)
    wb = openpyxl.load_workbook(fpath)

    # --- 표지 ---
    ws = wb['표지']
    ws.cell(row=11, column=4, value='2026.04.16')
    ws.cell(row=13, column=4, value='2.0')
    print("  [표지] version → 2.0, date → 2026.04.16")

    # --- 개정이력 ---
    ws = wb['개정이력']
    append_styled_row(ws, ['2.0', '2026.04.16', '소스코드 팩트 기반 최신화', '곽경훈'], style_row=4)
    print("  [개정이력] Added v2.0 row")

    # --- 기능구조도 ---
    ws = wb['기능구조도']

    # Fix typo: 인체관리 → 인증관리 (rows 4 and 5)
    fix_count = 0
    for r in range(4, ws.max_row + 1):
        cell = ws.cell(row=r, column=2)
        if cell.value == '인체관리':
            cell.value = '인증관리'
            fix_count += 1
    print(f"  [기능구조도] Fixed '인체관리' → '인증관리' in {fix_count} rows")

    # Update header date
    ws.cell(row=1, column=8, value='2026.04.16')

    style_r = 17  # last existing data row

    new_rows = [
        ['FC-PDF-015', '일반사용자', '홈페이지', '', '',
         ' -. 비로그인 사용자 접근 가능\n -. 서비스 소개 및 기능 안내\n -. 로그인/회원가입 유도\n -. URL: /home (기본 / → /home 리다이렉트)',
         '신규', ''],

        ['FC-PDF-016', '일반사용자', 'AI 챗봇', 'Omega-Prime RAG', '',
         ' -. POST /panel/chat: 인증 필수 실시간 AI 대화\n -. 하이브리드 RAG: Vector(0.4) + BM25(0.25) + Meta(0.35)\n -. CrossEncoder 리랭킹 (bge-reranker-v2-m3-ko)\n -. 4개 에이전트: Router→Planner→Judge→Synthesizer\n -. 한국어 금융 공시 특화 분석',
         '신규', ''],

        ['FC-PDF-017', '일반사용자', 'AI 챗봇', 'DART 공시 검색', '',
         ' -. POST /panel/search: 종목명/코드 기반 공시 검색\n -. GET /panel/autocomplete: 상장법인 자동완성 (80,000건)\n -. corpCode.xml 24시간 캐싱\n -. DART 공시 URL 직접 링크',
         '신규', ''],

        ['FC-PDF-018', '일반사용자', '사이드 패널', '실시간 모니터링', '',
         ' -. GET /panel/stats: DB 실시간 집계 통계\n -. GET /panel/system-status: 서비스 헬스체크 + 경보 레벨 (ok/warning/critical)\n -. GET /panel/activity-log: 최근 활동 로그',
         '신규', ''],

        ['FC-PDF-019', '인증관리', '회원탈퇴', '', '',
         ' -. POST /auth/request-withdraw: 비밀번호 + \'탈퇴합니다\' 확인문구 검증\n -. 15분 유효 이메일 인증 토큰 발송\n -. POST /auth/confirm-withdraw: PII 익명화 + 비활성화\n -. PIPA 준수: 이메일/사용자명/비밀번호 해시 즉시 익명화',
         '신규', ''],

        ['FC-PDF-020', '인증관리', '비밀번호 변경 (이메일 인증)', '', '',
         ' -. POST /auth/request-password-change: 새 비밀번호 해시를 토큰에 포함\n -. 15분 유효 인증 메일 발송\n -. POST /auth/confirm-password-change: 토큰 검증 후 비밀번호 적용',
         '신규', ''],

        ['FC-PDF-021', '일반사용자', '문서 관리', 'PDF 보고서', '',
         ' -. GET /documents/download-report/{id}: PDF 다운로드\n -. GET /documents/preview-report/{id}: 브라우저 미리보기\n -. fpdf2 기반 자동 생성',
         '신규', ''],

        ['FC-PDF-022', '관리자', '벡터 인덱스 관리', '', '',
         ' -. POST /panel/vector/rebuild: ChromaDB 전체 재인덱싱 (백그라운드)\n -. GET /panel/vector/stats: 인덱스 현황 (컬렉션명, 벡터수, 모델)',
         '신규', 'admin 권한만 접근'],
    ]

    for row_data in new_rows:
        append_styled_row(ws, row_data, style_row=style_r)

    print(f"  [기능구조도] Added {len(new_rows)} new rows (FC-PDF-015 ~ FC-PDF-022)")

    wb.save(fpath)
    print(f"  [SAVED] {fname}")


# ═══════════════════════════════════════════════════════════════════════════
# FILE 3: 03_설계_화면정의서_v1.0.xlsx
# ═══════════════════════════════════════════════════════════════════════════
def update_file_03():
    fname = '03_설계_화면정의서_v1.0.xlsx'
    fpath = os.path.join(BASE_DIR, fname)
    print(f"\n[FILE 3] {fname}")
    backup_file(fpath)
    wb = openpyxl.load_workbook(fpath)

    # --- 표지 ---
    ws = wb['표지']
    ws.cell(row=11, column=4, value='2026.04.16')
    ws.cell(row=13, column=4, value='2.0')
    ws.cell(row=14, column=4, value='16개')
    print("  [표지] version → 2.0, date → 2026.04.16, 총 화면 수 → 16개")

    # --- 개정이력 ---
    ws = wb['개정이력']
    append_styled_row(ws, ['2.0', '2026.04.16', 'URL 오류 수정, 홈페이지/회원탈퇴 화면 추가', '곽경훈'], style_row=4)
    print("  [개정이력] Added v2.0 row")

    # --- 화면목록 ---
    ws = wb['화면목록']

    # Fix URL for row 6 (SCR-003, row index 6 = data row 3): / → /upload
    ws.cell(row=6, column=4, value='/upload')
    print("  [화면목록] Fixed SCR-003 URL: / → /upload")

    # Fix URL for row 10 (SCR-006, row index 10 = data row 7): /admin → /admin/dashboard
    ws.cell(row=10, column=4, value='/admin/dashboard')
    print("  [화면목록] Fixed SCR-006 URL: /admin → /admin/dashboard")

    # Add new rows
    style_r = 17  # last data row
    # Row 15: SCR-014
    r15 = append_styled_row(ws, [15, 'SCR-014', '홈페이지', '/home', '전체',
                                  '서비스 소개 (비로그인 접근 가능, 기본 라우트)'], style_row=style_r)
    # Row 16: SCR-015
    r16 = append_styled_row(ws, [16, 'SCR-015', '회원탈퇴 인증', '/verify-withdraw', '전체',
                                  '회원탈퇴 이메일 인증 토큰 처리'], style_row=style_r)
    print(f"  [화면목록] Added SCR-014 (row {r15}), SCR-015 (row {r16})")

    wb.save(fpath)
    print(f"  [SAVED] {fname}")


# ═══════════════════════════════════════════════════════════════════════════
# FILE 4: 04_설계_ERD_v1.0.xlsx
# ═══════════════════════════════════════════════════════════════════════════
def update_file_04():
    fname = '04_설계_ERD_v1.0.xlsx'
    fpath = os.path.join(BASE_DIR, fname)
    print(f"\n[FILE 4] {fname}")
    backup_file(fpath)
    wb = openpyxl.load_workbook(fpath)

    # --- 표지 ---
    ws = wb['표지']
    ws.cell(row=11, column=4, value='2026.04.16')
    ws.cell(row=13, column=4, value='2.0')
    print("  [표지] version → 2.0, date → 2026.04.16")

    # --- 개정이력: need to CREATE this sheet (does not exist) ---
    if '개정이력' not in wb.sheetnames:
        ws = wb.create_sheet('개정이력', 1)  # insert after 표지
        # Build header structure matching other files
        ws.cell(row=1, column=1, value='문서 개정 이력표')
        ws.cell(row=1, column=1).font = Font(name='맑은 고딕', size=12, bold=True)
        # Row 2 blank
        headers = ['버전', '날짜', '내용', '작성자']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = Font(name='맑은 고딕', size=10, bold=True)
            cell.alignment = ALIGN_CC_WRAP
            cell.border = THIN_BORDER

        # v1.0 row
        v1_data = ['1.0', '2026.04.13', '최초 제정', '곽경훈']
        for c, v in enumerate(v1_data, 1):
            cell = ws.cell(row=4, column=c, value=v)
            cell.font = FONT_10
            cell.alignment = ALIGN_CC_WRAP
            cell.border = THIN_BORDER

        # v2.0 row
        v2_data = ['2.0', '2026.04.16', '소스코드 대조 검증 완료 (변경 없음)', '곽경훈']
        for c, v in enumerate(v2_data, 1):
            cell = ws.cell(row=5, column=c, value=v)
            cell.font = FONT_10
            cell.alignment = ALIGN_CC_WRAP
            cell.border = THIN_BORDER

        print("  [개정이력] Created sheet with v1.0 + v2.0 rows")
    else:
        ws = wb['개정이력']
        append_styled_row(ws, ['2.0', '2026.04.16', '소스코드 대조 검증 완료 (변경 없음)', '곽경훈'], style_row=4)
        print("  [개정이력] Added v2.0 row")

    wb.save(fpath)
    print(f"  [SAVED] {fname}")


# ═══════════════════════════════════════════════════════════════════════════
# FILE 5: 05_설계_프로그램목록_v1.0.xlsx
# ═══════════════════════════════════════════════════════════════════════════
def update_file_05():
    fname = '05_설계_프로그램목록_v1.0.xlsx'
    fpath = os.path.join(BASE_DIR, fname)
    print(f"\n[FILE 5] {fname}")
    backup_file(fpath)
    wb = openpyxl.load_workbook(fpath)

    # --- 표지 ---
    ws = wb['표지']
    ws.cell(row=11, column=4, value='2026.04.16')
    ws.cell(row=13, column=4, value='2.0')
    print("  [표지] version → 2.0, date → 2026.04.16")

    # --- 프론트엔드 ---
    ws = wb['프론트엔드']
    style_r = ws.max_row  # 18

    fe_rows = [
        ['PG-FE-018', '화면', '홈페이지', 'SCR-014', '/home',
         '서비스 소개 홈페이지 (비로그인 접근 가능)', '곽경훈 (4조)'],
        ['PG-FE-019', '화면', '회원탈퇴 인증 페이지', 'SCR-015', '/verify-withdraw',
         '회원탈퇴 이메일 인증 토큰 처리', '곽경훈 (4조)'],
        ['PG-FE-020', '컴포넌트', 'AI 챗봇 (ChatBot)', '-', '(사이드 패널)',
         'Omega-Prime RAG 챗봇 UI + DART 검색', '곽경훈 (4조)'],
        ['PG-FE-021', '컴포넌트', '사이드 패널 (SideDecorations)', '-', '(공통)',
         '좌/우 사이드 패널: 통계, 시스템 상태, 활동 로그', '곽경훈 (4조)'],
    ]

    for row_data in fe_rows:
        append_styled_row(ws, row_data, style_row=style_r)
    print(f"  [프론트엔드] Added {len(fe_rows)} rows (PG-FE-018 ~ PG-FE-021)")

    # --- 백엔드 API ---
    ws = wb['백엔드 API']
    style_r = ws.max_row  # 43

    be_rows = [
        ['PG-BE-043', '인증', '회원탈퇴 요청 API', 'POST', '/auth/request-withdraw',
         '비밀번호+확인문구 검증 후 탈퇴 인증 메일 발송 (15분 토큰)', '곽경훈 (4조)'],
        ['PG-BE-044', '인증', '회원탈퇴 확인 API', 'POST', '/auth/confirm-withdraw',
         '토큰 검증 후 PII 익명화 + 계정 비활성화 (PIPA 준수)', '곽경훈 (4조)'],
    ]

    for row_data in be_rows:
        append_styled_row(ws, row_data, style_row=style_r)
    print(f"  [백엔드 API] Added {len(be_rows)} rows (PG-BE-043 ~ PG-BE-044)")

    wb.save(fpath)
    print(f"  [SAVED] {fname}")


# ═══════════════════════════════════════════════════════════════════════════
# FILES 6-10: Only update 개정이력 (or create if missing)
# ═══════════════════════════════════════════════════════════════════════════
def update_revision_only(fname, file_num):
    """Add 2.0 revision entry. Create 개정이력 sheet if absent."""
    fpath = os.path.join(BASE_DIR, fname)
    print(f"\n[FILE {file_num}] {fname}")
    backup_file(fpath)
    wb = openpyxl.load_workbook(fpath)

    v2_data = ['2.0', '2026.04.16', '소스코드 기반 검증 완료', '곽경훈']

    if '개정이력' in wb.sheetnames:
        ws = wb['개정이력']
        append_styled_row(ws, v2_data, style_row=4)
        print("  [개정이력] Added v2.0 row")
    else:
        # Create 개정이력 sheet (for files 08, 09, 10 which lack it)
        ws = wb.create_sheet('개정이력', 1)  # after 표지
        ws.cell(row=1, column=1, value='문서 개정 이력표')
        ws.cell(row=1, column=1).font = Font(name='맑은 고딕', size=12, bold=True)

        headers = ['버전', '날짜', '내용', '작성자']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = Font(name='맑은 고딕', size=10, bold=True)
            cell.alignment = ALIGN_CC_WRAP
            cell.border = THIN_BORDER

        v1_data = ['1.0', '2026.04.13', '최초 제정', '곽경훈']
        for c, v in enumerate(v1_data, 1):
            cell = ws.cell(row=4, column=c, value=v)
            cell.font = FONT_10
            cell.alignment = ALIGN_CC_WRAP
            cell.border = THIN_BORDER

        for c, v in enumerate(v2_data, 1):
            cell = ws.cell(row=5, column=c, value=v)
            cell.font = FONT_10
            cell.alignment = ALIGN_CC_WRAP
            cell.border = THIN_BORDER

        print("  [개정이력] Created sheet with v1.0 + v2.0 rows")

    wb.save(fpath)
    print(f"  [SAVED] {fname}")


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    print("=" * 70)
    print("Omega CivicFlow v4 - Deliverables Update Script")
    print("Target: 2026.04.16 factual corrections")
    print("=" * 70)

    update_file_01()
    update_file_02()
    update_file_03()
    update_file_04()
    update_file_05()

    update_revision_only('06_테스트_단위테스트결과서_v1.0.xlsx', 6)
    update_revision_only('07_테스트_통합테스트시나리오결과_v1.0.xlsx', 7)
    update_revision_only('08_이행_사용자매뉴얼_v1.0.xlsx', 8)
    update_revision_only('09_이행_운영자매뉴얼_v1.0.xlsx', 9)
    update_revision_only('10_이행_LLM구동절차_및_설정가이드_v1.0.xlsx', 10)

    print("\n" + "=" * 70)
    print("ALL 10 FILES UPDATED SUCCESSFULLY")
    print("=" * 70)


if __name__ == '__main__':
    main()
