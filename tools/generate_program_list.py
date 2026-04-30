"""Omega CivicFlow v4 프로그램목록 Excel 생성 스크립트"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

thin = Side(style="thin")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
data_font = Font(name="맑은 고딕", size=10)
c_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
l_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


def hdr_row(ws, row, n):
    for c in range(1, n + 1):
        cl = ws.cell(row=row, column=c)
        cl.font = hdr_font
        cl.fill = hdr_fill
        cl.alignment = c_align
        cl.border = border_all


def data_row(ws, row, n):
    for c in range(1, n + 1):
        cl = ws.cell(row=row, column=c)
        cl.font = data_font
        cl.alignment = l_align if c > 3 else c_align
        cl.border = border_all


# ═══════════════════════════════════════════════════
# Sheet 1: 표지
# ═══════════════════════════════════════════════════
ws = wb.active
ws.title = "표지"
ws.merge_cells("B4:H4")
ws["B4"] = "한국 금융 공시 분석 플랫폼"
ws["B4"].font = Font(name="맑은 고딕", bold=True, size=14)
ws["B4"].alignment = c_align

ws.merge_cells("B6:H7")
ws["B6"] = "Omega CivicFlow v4"
ws["B6"].font = Font(name="맑은 고딕", bold=True, size=22, color="2F5496")
ws["B6"].alignment = c_align

ws.merge_cells("B9:H9")
ws["B9"] = "프 로 그 램 목 록"
ws["B9"].font = Font(name="맑은 고딕", bold=True, size=18, color="2F5496")
ws["B9"].alignment = c_align

meta = [
    ("시스템 명칭", "Omega CivicFlow v4"),
    ("문서 번호", "CIV-D-0100"),
    ("버전", "Ver 4.0"),
    ("작성일", datetime.now().strftime("%Y-%m-%d")),
    ("작성자", "hibou"),
    ("기술 스택", "FastAPI + React + ChromaDB + EXAONE 3.5 + BGE-M3"),
]
for i, (k, v) in enumerate(meta):
    r = 12 + i
    ws.merge_cells(f"C{r}:D{r}")
    ws[f"C{r}"] = k
    ws[f"C{r}"].font = Font(name="맑은 고딕", bold=True, size=11)
    ws[f"C{r}"].alignment = c_align
    ws[f"C{r}"].fill = sub_fill
    ws[f"C{r}"].border = border_all
    ws[f"D{r}"].border = border_all
    ws.merge_cells(f"E{r}:G{r}")
    ws[f"E{r}"] = v
    ws[f"E{r}"].font = data_font
    ws[f"E{r}"].alignment = c_align
    ws[f"E{r}"].border = border_all
    ws[f"F{r}"].border = border_all
    ws[f"G{r}"].border = border_all
for c in range(1, 9):
    ws.column_dimensions[get_column_letter(c)].width = 14

# ═══════════════════════════════════════════════════
# Sheet 2: 변경이력
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("변경이력")
ws2.merge_cells("B2:F2")
ws2["B2"] = "문서 변경 이력"
ws2["B2"].font = Font(name="맑은 고딕", bold=True, size=14)

for c, h in enumerate(["버전", "변경일자", "변경내용", "작성자", "승인자"], 2):
    ws2.cell(row=4, column=c, value=h)
    ws2.cell(row=4, column=c).font = hdr_font
    ws2.cell(row=4, column=c).fill = hdr_fill
    ws2.cell(row=4, column=c).alignment = c_align
    ws2.cell(row=4, column=c).border = border_all

hist = [
    ("1.0", "2025-03-15", "최초 작성 - 기본 아키텍처 설계", "hibou", "hibou"),
    ("2.0", "2025-03-27", "OCR + LLM 분석 파이프라인 구현", "hibou", "hibou"),
    ("3.0", "2025-04-06", "Phase 3: BGE-M3 임베딩 + ChromaDB RAG 통합", "hibou", "hibou"),
    ("4.0", "2025-04-10", "Phase 4: Omega-Prime 챗봇 에이전트 완성", "hibou", "hibou"),
    ("4.1", "2025-04-12", "RAGAS 95.8/100 달성, 프로그램목록 정리", "hibou", "hibou"),
]
for i, row_data in enumerate(hist):
    for c, v in enumerate(row_data, 2):
        ws2.cell(row=5 + i, column=c, value=v)
        ws2.cell(row=5 + i, column=c).font = data_font
        ws2.cell(row=5 + i, column=c).alignment = c_align
        ws2.cell(row=5 + i, column=c).border = border_all

ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 45
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 12

# ═══════════════════════════════════════════════════
# Sheet 3: Backend 프로그램목록
# ═══════════════════════════════════════════════════
ws3 = wb.create_sheet("Backend 프로그램목록")
cols3 = ["No", "대분류", "중분류", "모듈/기능", "프로그램ID", "상태", "프로그램 설명", "CRUD", "의존 기술", "비고"]
for c, h in enumerate(cols3, 1):
    ws3.cell(row=1, column=c, value=h)
hdr_row(ws3, 1, 10)

be = [
    # Backend Core
    (1, "Backend Core", "앱 서버", "앱 초기화", "main.py", "기존",
     "FastAPI 앱 생명주기, CORS, 라우터 등록, ChromaDB 연결", "-", "FastAPI, Uvicorn", "진입점"),
    (2, "Backend Core", "설정", "환경변수", "config.py", "기존",
     "전역 설정 및 .env 환경변수 관리", "R", "pydantic-settings", ""),
    (3, "Backend Core", "DB", "ORM 세션", "database.py", "기존",
     "SQLAlchemy 엔진 생성 및 세션 의존성 주입", "R", "SQLAlchemy", ""),
    # AI Agent
    (4, "AI Agent", "오케스트레이터", "추론 파이프라인", "agents/orchestrator.py", "변경",
     "멀티스테이지 추론: Router > Planner > Judge > Synthesizer > Critic > Reviser", "R", "Ollama, EXAONE", "핵심"),
    (5, "AI Agent", "LLM 클라이언트", "LLM 호출", "agents/llm_client.py", "변경",
     "Ollama/vLLM LLM 클라이언트, JSON 파싱, 폴백 관리", "R", "httpx, Ollama", ""),
    (6, "AI Agent", "프롬프트", "프롬프트 빌더", "agents/prompts.py", "변경",
     "Omega-Prime 시스템 프롬프트 및 6단계 프롬프트 템플릿", "-", "-", ""),
    (7, "AI Agent", "스키마", "데이터 모델", "agents/schemas.py", "기존",
     "RouterResult, PlanResult, JudgeResult Pydantic 스키마", "-", "Pydantic v2", ""),
    # API Router
    (8, "API Router", "인증", "회원 인증", "routers/auth.py", "기존",
     "회원가입, 로그인, 이메일인증, 비밀번호 재설정 API", "C,R,U", "JWT, bcrypt", ""),
    (9, "API Router", "문서", "문서 관리", "routers/documents.py", "기존",
     "문서 업로드, OCR, LLM 분석, PDF 보고서 생성 API", "C,R,U,D", "FastAPI", ""),
    (10, "API Router", "관리자", "관리자 기능", "routers/admin.py", "기존",
     "관리자 대시보드, 회원관리, 문서재분류 API", "R,U,D", "FastAPI", ""),
    (11, "API Router", "패널", "통합 인텔리전스", "routers/panel.py", "변경",
     "시스템통계, 활동로그, DART검색, 챗봇 SSE API", "C,R", "FastAPI, SSE", "챗봇 EP"),
    # Service - 검색/RAG
    (12, "Service", "인지 검색", "하이브리드 검색", "services/cognitive_search_safe.py", "변경",
     "벡터+BM25+퍼지+CrossEncoder 리랭킹 하이브리드 검색", "R", "ChromaDB, CE", "핵심"),
    (13, "Service", "벡터 DB", "ChromaDB", "services/vector_service.py", "변경",
     "ChromaDB 하이브리드검색, 리랭킹, 엔트로피 소각", "C,R,D", "ChromaDB, BGE-M3", ""),
    (14, "Service", "RAG 검색", "쿼리 정제", "services/agent_retrieval.py", "신규",
     "하이브리드 RAG 쿼리 정제 및 다중 검색 전략", "R", "sentence-transformers", "Phase 4"),
    (15, "Service", "지식 조회", "구조화 SQL", "services/chat_knowledge_service.py", "신규",
     "회사요약, 재무팩트, 순위비교, 추세 SQL 조회", "R", "SQLAlchemy", "Phase 4"),
    # Service - Agent
    (16, "Service", "챗봇 에이전트", "안전 실행", "services/chat_agent_safe_service.py", "신규",
     "회사명 바인딩, 의도분류, RAG/SQL 라우팅, 에러처리", "R", "EXAONE, ChromaDB", "Phase 4 핵심"),
    (17, "Service", "챗봇 래퍼", "에이전트 위임", "services/chat_agent_service.py", "기존",
     "챗봇 에이전트 실행 래퍼 (safe 버전 위임)", "R", "-", ""),
    (18, "Service", "세션 메모리", "대화 컨텍스트", "services/agent_memory.py", "신규",
     "슬롯 기반 대화 이력 요약 및 세션 컨텍스트 빌더", "C,R", "-", "Phase 4"),
    (19, "Service", "챗봇 프로필", "설정 관리", "services/chat_profile_service.py", "기존",
     "챗봇 프로필 및 설정 JSON 관리", "C,R,U", "-", ""),
    # Service - 문서처리
    (20, "Service", "OCR", "텍스트 추출", "services/ocr_service.py", "기존",
     "EasyOCR 다중언어(한/영) 텍스트 추출 및 품질 최적화", "R", "EasyOCR", ""),
    (21, "Service", "LLM 분석", "문서 분석", "services/llm_service.py", "기존",
     "Ollama 기반 문서 분류, 요약, 핵심 추출 엔진", "R", "Ollama, EXAONE", ""),
    (22, "Service", "메타데이터 추출", "전처리", "services/document_metadata_extractor.py", "기존",
     "OCR 텍스트에서 회사명, 섹션 사전 추출", "R", "-", ""),
    (23, "Service", "DART 파싱", "XBRL 파서", "services/dart_file_parser.py", "기존",
     "XBRL ZIP, XLS 파일 파싱 및 텍스트 추출", "R", "xlrd, zipfile", ""),
    (24, "Service", "코드 추출", "규칙 기반", "services/code_only_extractor.py", "기존",
     "정규식/규칙 기반 DART 구조데이터 추출", "R", "re", ""),
    # Service - 텍스트
    (25, "Service", "텍스트 전처리", "표/섹션 복원", "services/text_preprocessor.py", "기존",
     "표 구조 복원, 섹션 태깅, 숫자 정규화", "R", "-", ""),
    (26, "Service", "텍스트 품질", "품질 검증", "services/text_quality.py", "기존",
     "OCR 텍스트 품질 검증 및 가독성 점수 측정", "R", "-", ""),
    (27, "Service", "텍스트 요약", "TextRank", "services/text_summarizer.py", "기존",
     "TextRank + 템플릿 기반 문서 요약 분석기", "R", "-", ""),
    (28, "Service", "서사 요약", "자연어 생성", "services/narrative_summarizer.py", "기존",
     "템플릿 기반 자연어 문서 요약 생성기", "R", "-", ""),
    # Service - 기타
    (29, "Service", "회사명 정규화", "별칭 마스터", "services/company_alias_master.py", "변경",
     "회사명 정규화 및 별칭 매핑 마스터 사전", "R", "-", "1,000+기업"),
    (30, "Service", "종목명 정규화", "음독 변환", "services/stock_name_normalizer.py", "기존",
     "음독에서 영문 회사명 정규화 변환기", "R", "-", ""),
    (31, "Service", "인증", "JWT 관리", "services/auth_service.py", "기존",
     "JWT, bcrypt, 이메일토큰 인증 및 인가", "C,R,U", "PyJWT, bcrypt", ""),
    (32, "Service", "이메일", "SMTP 발송", "services/email_service.py", "기존",
     "Gmail SMTP 이메일 발송 및 첨부파일 처리", "C", "smtplib", ""),
    (33, "Service", "PDF 리포트", "보고서 생성", "services/pdf_report_service.py", "기존",
     "A4 구조화 PDF 분석 보고서 생성 엔진", "C", "FPDF2", ""),
    (34, "Service", "인사이트", "Gemini 분석", "services/insight_service.py", "기존",
     "Gemini 2.5 Pro 전략적 인사이트 도출", "C,R", "Google GenAI", ""),
    (35, "Service", "Supervisor", "사후 검증", "services/omega_supervisor.py", "기존",
     "인사이트 Omega-Prime 사후 검증 및 보강", "R,U", "Gemini", ""),
    (36, "Service", "임베딩 전략", "청크 설계", "services/embedding_strategy.py", "기존",
     "재무문서 청크 전략 및 임베딩 템플릿", "R", "BGE-M3", ""),
    (37, "Service", "메타 검증", "앵커링", "services/metadata_validator.py", "기존",
     "메타데이터 앵커링, 검증, 안전 렌더링", "R", "-", ""),
    (38, "Service", "VLM", "비전 LLM", "services/vlm_service.py", "기존",
     "Qwen2.5-VL 파인튜닝 모델 vLLM 연동", "R", "vLLM, Qwen2.5-VL", ""),
    # Data Model
    (39, "Data Model", "ORM", "DB 테이블", "models/models.py", "기존",
     "User, Document, Page, OCR, Analysis ORM 모델", "-", "SQLAlchemy", ""),
    (40, "Data Model", "API 스키마", "요청/응답", "schemas/schemas.py", "기존",
     "Auth, Document, Admin Pydantic 스키마", "-", "Pydantic v2", ""),
    (41, "Data Model", "Supervisor", "입출력 모델", "schemas/supervisor_schemas.py", "기존",
     "Supervisor 입출력 Pydantic v2 스키마", "-", "Pydantic v2", ""),
    # Worker
    (42, "Worker", "Celery", "작업 큐", "workers/celery_app.py", "기존",
     "Redis 기반 Celery 비동기 작업 큐 설정", "-", "Celery, Redis", ""),
    (43, "Worker", "디스패처", "태스크 발행", "workers/task_dispatcher.py", "기존",
     "Redis 직접 메시지 발행 태스크 디스패처", "C", "Redis", ""),
    (44, "Worker", "태스크", "OCR+분석", "workers/tasks.py", "기존",
     "OCR + LLM 분석 백그라운드 태스크 정의", "C,R,U", "Celery", ""),
]

for i, row in enumerate(be):
    r = 2 + i
    for c, v in enumerate(row):
        ws3.cell(row=r, column=c + 1, value=v)
    data_row(ws3, r, 10)

widths3 = [5, 13, 14, 15, 38, 5, 52, 9, 20, 12]
for c, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

cat_c = {
    "Backend Core": "E2EFDA", "AI Agent": "FCE4D6", "API Router": "D9E2F3",
    "Service": "FFF2CC", "Data Model": "E2D9F3", "Worker": "D6E4F0",
}
for r in range(2, 2 + len(be)):
    cat = ws3.cell(row=r, column=2).value
    if cat in cat_c:
        ws3.cell(row=r, column=2).fill = PatternFill(
            start_color=cat_c[cat], end_color=cat_c[cat], fill_type="solid"
        )
ws3.freeze_panes = "A2"

# ═══════════════════════════════════════════════════
# Sheet 4: Frontend 프로그램목록
# ═══════════════════════════════════════════════════
ws4 = wb.create_sheet("Frontend 프로그램목록")
cols4 = ["No", "대분류", "중분류", "모듈/기능", "프로그램ID", "상태", "프로그램 설명", "화면 경로", "CRUD", "비고"]
for c, h in enumerate(cols4, 1):
    ws4.cell(row=1, column=c, value=h)
hdr_row(ws4, 1, 10)

fe = [
    (1, "Core", "진입점", "React 부트스트랩", "src/main.jsx", "기존",
     "React DOM 렌더링 진입점", "-", "-", ""),
    (2, "Core", "라우팅", "앱 라우터", "src/App.jsx", "기존",
     "React Router 라우팅 및 레이아웃", "/*", "R", ""),
    (3, "Component", "챗봇", "AI 대화 UI", "src/components/ChatBot.jsx", "변경",
     "Omega-Prime AI 챗봇 SSE 스트리밍 대화", "-", "C,R", "Phase 4"),
    (4, "Component", "네비게이션", "메뉴 바", "src/components/Navbar.jsx", "기존",
     "상단 네비게이션 바 및 메뉴", "-", "R", ""),
    (5, "Component", "라우트 보호", "인증 가드", "src/components/ProtectedRoute.jsx", "기존",
     "인증 기반 라우트 접근 제어", "-", "R", ""),
    (6, "Component", "사이드 패널", "실시간 통계", "src/components/SideDecorations.jsx", "변경",
     "사이드 실시간 통계, 시스템상태, 활동로그", "-", "R", ""),
    (7, "Context", "인증 상태", "전역 상태", "src/contexts/AuthContext.jsx", "기존",
     "전역 인증 상태 React Context", "-", "C,R,U", ""),
    (8, "Page", "홈", "랜딩 페이지", "src/pages/HomePage.jsx", "기존",
     "공개 홈페이지 및 제품 소개", "/", "R", ""),
    (9, "Page", "로그인", "로그인", "src/pages/LoginPage.jsx", "기존",
     "사용자 로그인 폼", "/login", "R", ""),
    (10, "Page", "회원가입", "일반 가입", "src/pages/RegisterPage.jsx", "기존",
     "일반 사용자 회원가입", "/register", "C", ""),
    (11, "Page", "관리자 가입", "마스터 가입", "src/pages/AdminRegisterPage.jsx", "기존",
     "마스터키 관리자 계정 생성", "/admin/register", "C", ""),
    (12, "Page", "마이페이지", "개인 관리", "src/pages/MyPage.jsx", "기존",
     "개인 문서 관리, 프로필 편집, 챗봇", "/my", "C,R,U,D", ""),
    (13, "Page", "문서 업로드", "ZIP 업로드", "src/pages/UploadPage.jsx", "기존",
     "일괄 문서 업로드 및 ZIP 처리", "/upload", "C", ""),
    (14, "Page", "문서 상세", "분석 조회", "src/pages/DocumentDetail.jsx", "기존",
     "문서 상세 분석, 인사이트, PDF 미리보기", "/documents/:id", "R,U", ""),
    (15, "Page", "관리자 대시보드", "통합 관리", "src/pages/AdminDashboard.jsx", "기존",
     "관리자 대시보드 허브", "/admin", "R", ""),
    (16, "Page", "관리자 문서", "문서 관리", "src/pages/AdminDocuments.jsx", "기존",
     "전체 문서 관리 및 재분류", "/admin/documents", "R,U,D", ""),
    (17, "Page", "관리자 회원", "회원 관리", "src/pages/AdminUsers.jsx", "기존",
     "회원 역할, 활성화 관리", "/admin/users", "R,U", ""),
    (18, "Page", "비밀번호 찾기", "찾기 요청", "src/pages/ForgotPassword.jsx", "기존",
     "비밀번호 찾기 이메일 발송", "/forgot-password", "C", ""),
    (19, "Page", "비밀번호 재설정", "토큰 재설정", "src/pages/ResetPassword.jsx", "기존",
     "토큰 기반 비밀번호 재설정", "/reset-password/:token", "U", ""),
    (20, "Page", "이메일 인증", "인증 처리", "src/pages/VerifyEmail.jsx", "기존",
     "이메일 인증 토큰 검증", "/verify-email/:token", "U", ""),
    (21, "Page", "비밀번호 변경", "변경 검증", "src/pages/VerifyPasswordChange.jsx", "기존",
     "비밀번호 변경 인증 검증", "/verify-pw-change/:token", "U", ""),
]

for i, row in enumerate(fe):
    r = 2 + i
    for c, v in enumerate(row):
        ws4.cell(row=r, column=c + 1, value=v)
    data_row(ws4, r, 10)

widths4 = [5, 12, 14, 15, 40, 5, 45, 26, 9, 12]
for c, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(c)].width = w

fe_c = {"Core": "E2EFDA", "Component": "FCE4D6", "Context": "D9E2F3", "Page": "FFF2CC"}
for r in range(2, 2 + len(fe)):
    cat = ws4.cell(row=r, column=2).value
    if cat in fe_c:
        ws4.cell(row=r, column=2).fill = PatternFill(
            start_color=fe_c[cat], end_color=fe_c[cat], fill_type="solid"
        )
ws4.freeze_panes = "A2"

# ═══════════════════════════════════════════════════
# Sheet 5: Tools & 배치도구
# ═══════════════════════════════════════════════════
ws5 = wb.create_sheet("Tools 배치도구")
cols5 = ["No", "대분류", "중분류", "프로그램ID", "상태", "프로그램 설명", "실행환경", "비고"]
for c, h in enumerate(cols5, 1):
    ws5.cell(row=1, column=c, value=h)
hdr_row(ws5, 1, 8)

tools = [
    (1, "Backend Tool", "평가", "backend/tools/ragas_eval.py", "변경",
     "RAGAS 6-metric RAG 성능 평가 (95.8/100 달성)", "Local", "52 QA"),
    (2, "Backend Tool", "DART 배치", "backend/tools/dart_batch_pipeline.py", "기존",
     "XBRL ZIP에서 OCR, 임베딩, ChromaDB 일괄 적재", "Local", ""),
    (3, "Backend Tool", "XML 인제스트", "backend/tools/dart_xml_batch_ingest.py", "기존",
     "DART XML 배치 파싱 및 DB 인제스트", "Local", ""),
    (4, "Backend Tool", "재색인", "backend/tools/reindex_v2.py", "기존",
     "전체 문서 재색인 및 매니페스트 관리", "Local", ""),
    (5, "Backend Tool", "초기화", "backend/tools/clean_reset_pipeline.py", "기존",
     "DB + ChromaDB 완전 초기화 및 재구축", "Local", ""),
    (6, "Backend Tool", "회사링크 감사", "backend/tools/audit_company_links.py", "기존",
     "회사-문서 연결 무결성 감사", "Local", ""),
    (7, "Backend Tool", "팩트 백필", "backend/tools/backfill_facts_v2.py", "기존",
     "재무팩트 데이터 일괄 백필", "Local", ""),
    (8, "Backend Tool", "중국어 정화", "backend/tools/purge_chinese_chunks.py", "기존",
     "중국어 혼입 청크 탐지 및 제거", "Local", ""),
    (9, "Backend Tool", "QA 품질", "backend/tools/qa_narrative_quality.py", "기존",
     "QA 데이터셋 서사 품질 검증", "Local", ""),
    (10, "Backend Tool", "Supervisor 테스트", "backend/tools/simulate_supervisor.py", "기존",
     "Omega Supervisor 시뮬레이션", "Local", ""),
    (11, "Backend Tool", "디버그", "backend/tools/debug_trace.py", "기존",
     "검색 파이프라인 디버그 추적", "Local", ""),
    (12, "Backend Tool", "임베딩(A100)", "backend/tools/phase3_embedding_a100.py", "기존",
     "A100 GPU 전용 BGE-M3 배치 임베딩", "Colab", "284K chunks"),
    (13, "Backend Tool", "임베딩(RTX5070)", "backend/tools/phase3_embedding_rtx5070.py", "기존",
     "RTX 5070 로컬 GPU BGE-M3 배치 임베딩 (batch=16, 1024-dim)", "Local GPU", "RTX 5070"),
    (14, "Root Tool", "Colab A100", "tools/colab_a100_pipeline.py", "기존",
     "Colab A100 전체 DART 분석 파이프라인", "Colab A100", ""),
    (15, "Root Tool", "Colab H100", "tools/colab_h100_full_pipeline.py", "기존",
     "Colab H100 전체 파이프라인", "Colab H100", ""),
    (16, "Root Tool", "QLoRA 파인튜닝", "tools/dart_finetune_qlora.py", "기존",
     "Qwen2.5 QLoRA 파인튜닝 트레이닝 스크립트", "RunPod", ""),
    (17, "Root Tool", "PPTX 생성", "tools/generate_pptx.py", "기존",
     "포트폴리오 프레젠테이션 PPTX 자동 생성", "Local", ""),
    (18, "Root Tool", "ChromaDB 재구축", "tools/rebuild_chromadb.py", "기존",
     "ChromaDB 완전 초기화 및 재구축", "Local", ""),
    (19, "Root Tool", "Gemini 재분석", "tools/gemini_reanalyze_all.py", "기존",
     "Gemini 기반 전체 문서 재분석", "Cloud", ""),
    (20, "Root Tool", "PDF 재생성", "tools/regenerate_all_pdfs.py", "기존",
     "전체 PDF 보고서 일괄 재생성", "Local", ""),
    (21, "Root Tool", "데이터 준비", "tools/data_prep/*.py", "기존",
     "JSONL 변환, 중복제거, LabelStudio 태스크 (8개)", "Local", "학습 데이터"),
]

for i, row in enumerate(tools):
    r = 2 + i
    for c, v in enumerate(row):
        ws5.cell(row=r, column=c + 1, value=v)
    data_row(ws5, r, 8)

widths5 = [5, 14, 16, 42, 5, 50, 12, 14]
for c, w in enumerate(widths5, 1):
    ws5.column_dimensions[get_column_letter(c)].width = w

tc = {"Backend Tool": "FFF2CC", "Root Tool": "E2EFDA"}
for r in range(2, 2 + len(tools)):
    cat = ws5.cell(row=r, column=2).value
    if cat in tc:
        ws5.cell(row=r, column=2).fill = PatternFill(
            start_color=tc[cat], end_color=tc[cat], fill_type="solid"
        )
ws5.freeze_panes = "A2"

# ═══════════════════════════════════════════════════
# Sheet 6: Tests
# ═══════════════════════════════════════════════════
ws6 = wb.create_sheet("Tests")
cols6 = ["No", "테스트 파일", "설명", "대상 모듈", "비고"]
for c, h in enumerate(cols6, 1):
    ws6.cell(row=1, column=c, value=h)
hdr_row(ws6, 1, 5)

tests = [
    (1, "tests/test_chat_knowledge_service.py", "지식 서비스 단위 테스트", "chat_knowledge_service", ""),
    (2, "tests/test_chat_safe_retrieval.py", "안전 검색 통합 테스트", "chat_agent_safe_service", ""),
    (3, "tests/test_company_validation.py", "회사명 검증 테스트", "company_alias_master", ""),
    (4, "tests/test_deep_ranking_compare.py", "리랭킹 심층 비교 테스트", "cognitive_search_safe", ""),
    (5, "tests/test_integrity.py", "데이터 무결성 검증", "vector_service", ""),
    (6, "tests/test_metadata_extractor.py", "메타데이터 추출 테스트", "document_metadata_extractor", ""),
    (7, "tests/test_metadata_validator.py", "메타데이터 검증 테스트", "metadata_validator", ""),
    (8, "tests/test_reranker.py", "리랭커 성능 테스트", "cognitive_search_safe", ""),
    (9, "tests/test_text_quality.py", "텍스트 품질 검증 테스트", "text_quality", ""),
]

for i, row in enumerate(tests):
    r = 2 + i
    for c, v in enumerate(row):
        ws6.cell(row=r, column=c + 1, value=v)
    data_row(ws6, r, 5)

widths6 = [5, 42, 28, 28, 14]
for c, w in enumerate(widths6, 1):
    ws6.column_dimensions[get_column_letter(c)].width = w
ws6.freeze_panes = "A2"

# ═══════════════════════════════════════════════════
# Sheet 7: 통계 요약
# ═══════════════════════════════════════════════════
ws7 = wb.create_sheet("통계 요약")
ws7.merge_cells("B2:E2")
ws7["B2"] = "프로그램 통계 요약"
ws7["B2"].font = Font(name="맑은 고딕", bold=True, size=14)

for c, h in enumerate(["구분", "프로그램 수", "신규", "변경"], 2):
    ws7.cell(row=4, column=c, value=h)
    ws7.cell(row=4, column=c).font = hdr_font
    ws7.cell(row=4, column=c).fill = hdr_fill
    ws7.cell(row=4, column=c).alignment = c_align
    ws7.cell(row=4, column=c).border = border_all

stats = [
    ("Backend Core", 3, 0, 0),
    ("AI Agent", 4, 0, 3),
    ("API Router", 4, 0, 1),
    ("Service (검색/RAG)", 4, 2, 2),
    ("Service (Agent)", 4, 2, 0),
    ("Service (문서처리)", 5, 0, 0),
    ("Service (텍스트)", 4, 0, 0),
    ("Service (기타)", 10, 0, 1),
    ("Data Model", 3, 0, 0),
    ("Worker", 3, 0, 0),
    ("Frontend Component", 7, 0, 2),
    ("Frontend Page", 14, 0, 0),
    ("Backend Tool", 13, 0, 1),
    ("Root Tool", 8, 0, 0),
    ("Test", 9, 0, 0),
]

for i, (cat, tot, new, chg) in enumerate(stats):
    r = 5 + i
    for c, v in enumerate([cat, tot, new, chg], 2):
        ws7.cell(row=r, column=c, value=v)
        ws7.cell(row=r, column=c).font = data_font
        ws7.cell(row=r, column=c).alignment = c_align
        ws7.cell(row=r, column=c).border = border_all

r_t = 5 + len(stats)
for c, v in enumerate(
    ["합계", sum(s[1] for s in stats), sum(s[2] for s in stats), sum(s[3] for s in stats)], 2
):
    ws7.cell(row=r_t, column=c, value=v)
    ws7.cell(row=r_t, column=c).font = Font(name="맑은 고딕", bold=True, size=11)
    ws7.cell(row=r_t, column=c).alignment = c_align
    ws7.cell(row=r_t, column=c).fill = sub_fill
    ws7.cell(row=r_t, column=c).border = border_all

ws7.column_dimensions["B"].width = 22
ws7.column_dimensions["C"].width = 14
ws7.column_dimensions["D"].width = 10
ws7.column_dimensions["E"].width = 10

# ═══════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════
out = r"C:\Users\hibou\Omega_CivicFlow_v4\프로그램목록_Omega_CivicFlow_v4.xlsx"
wb.save(out)
print(f"OK: {out}")
print(f"Backend: {len(be)} | Frontend: {len(fe)} | Tools: {len(tools)} | Tests: {len(tests)}")
print(f"Total: {len(be) + len(fe) + len(tools) + len(tests)} programs")
print(f"Sheets: {wb.sheetnames}")
