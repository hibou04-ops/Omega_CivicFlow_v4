"""
산출물 11개 xlsx 파일의 팩트 정정 스크립트
- Insight 모델: gemini-2.5-flash → gemini-2.5-pro (Supervisor는 Flash 유지)
- 사용자 명시 사실: Insight=Pro, Supervisor(Omega Prime)=Flash
"""
import openpyxl
import os
import re

FOLDER = r"C:/Users/hibou/Downloads/산출물_4조_곽경훈"

# 정규식: "Insight" 문맥에서 gemini-2.5-flash를 gemini-2.5-pro로 치환
# Supervisor 단독 표기는 변경하지 않음

REPLACE_RULES = [
    # (pattern, replacement) - 정규식
    # Pattern 1: "gemini-2.5-flash — Vertex AI (Insight/Supervisor 엔진)"
    # → Insight는 Pro, Supervisor는 Flash로 분리 명시
    (
        r"gemini-2\.5-flash — Vertex AI \(Insight/Supervisor 엔진\)",
        "gemini-2.5-pro (Insight) + gemini-2.5-flash (Supervisor) — Vertex AI",
    ),
    # Pattern 2: "gemini-2.5-flash — Vertex AI (Insight/Supervisor)"
    (
        r"gemini-2\.5-flash — Vertex AI \(Insight/Supervisor\)",
        "gemini-2.5-pro (Insight) + gemini-2.5-flash (Supervisor)",
    ),
    # Pattern 3: "gemini-2.5-flash (Vertex AI) — Insight/Supervisor 엔진"
    (
        r"gemini-2\.5-flash \(Vertex AI\) — Insight/Supervisor 엔진",
        "gemini-2.5-pro (Insight) + gemini-2.5-flash (Supervisor) — Vertex AI",
    ),
    # Pattern 4: "gemini-2.5-flash (Vertex AI) Insight 엔진"
    (
        r"gemini-2\.5-flash \(Vertex AI\) Insight 엔진",
        "gemini-2.5-pro (Vertex AI) Insight 엔진",
    ),
    # Pattern 5: "gemini-2.5-flash 기반 전략 Insight 생성"
    (
        r"gemini-2\.5-flash 기반 전략 Insight",
        "gemini-2.5-pro 기반 전략 Insight",
    ),
    # Pattern 6: "gemini-2.5-flash Insight 생성"
    (
        r"gemini-2\.5-flash Insight 생성",
        "gemini-2.5-pro Insight 생성 + gemini-2.5-flash Supervisor 감독",
    ),
    # Pattern 7: "gemini-2.5-flash — Vertex AI (전략 Insight/Supervisor 엔진)"
    (
        r"gemini-2\.5-flash — Vertex AI \(전략 Insight/Supervisor 엔진\)",
        "gemini-2.5-pro (전략 Insight) + gemini-2.5-flash (Supervisor) — Vertex AI",
    ),
    # Pattern 8: "Insight 모델 | gemini-2.5-flash" 단독
    # 단독 표기는 컨텍스트 봐서: 라벨이 "Insight"면 Pro, "Supervisor"면 Flash
    # 이건 셀 단위로 하단에서 처리
]


def fix_workbook(path):
    """xlsx를 열어서 패턴 매칭으로 셀 내용 치환. 변경된 셀 수 반환."""
    wb = openpyxl.load_workbook(path)
    changes = 0

    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original = cell.value
                    new_val = original

                    # 일반 패턴 적용
                    for pattern, replacement in REPLACE_RULES:
                        new_val = re.sub(pattern, replacement, new_val)

                    # 컨텍스트 기반 단독 셀 치환
                    # "Insight" 라벨 셀 옆에 "gemini-2.5-flash" 단독 값이 있는 경우
                    # → cell의 인접 셀 (좌측)에 "Insight"가 있으면 변경
                    if new_val == original and new_val.strip() == "gemini-2.5-flash":
                        # 좌측 셀 확인
                        if cell.column > 1:
                            left = ws.cell(row=cell.row, column=cell.column - 1).value
                            if left and isinstance(left, str) and "insight" in left.lower() and "supervisor" not in left.lower():
                                new_val = "gemini-2.5-pro"

                    if new_val != original:
                        cell.value = new_val
                        changes += 1
                        print(f"  [{sn}] {original[:60]}... → {new_val[:60]}...")

    if changes > 0:
        wb.save(path)
    return changes


def main():
    files = [f for f in os.listdir(FOLDER) if f.endswith(".xlsx")]
    total_changes = 0

    for f in sorted(files):
        path = os.path.join(FOLDER, f)
        print(f"\n=== {f} ===")
        try:
            n = fix_workbook(path)
            print(f"  → {n} cells updated")
            total_changes += n
        except PermissionError:
            print(f"  ⚠ PermissionError — file is open in Excel")
        except Exception as e:
            print(f"  ⚠ Error: {e}")

    print(f"\n총 변경: {total_changes} 셀")


if __name__ == "__main__":
    main()
